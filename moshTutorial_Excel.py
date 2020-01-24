import openpyxl as xl

# First load your excel file through workbook
work_book = xl.load_workbook('./Python_Excel/transactions.xlsx')
# Access sheet by name (Case sensitive)
sheet = work_book['Sheet1']
# Index sheet with cell co-ords or use special method cell which demands actual numeral co-ords.
# check value with value method on returned value
cell = sheet['A1']
print(cell.value)
cell = sheet.cell(2, 3)
print(cell.value)

# How many rows are are in the sheet
max = sheet.max_row
sheet.cell(1, 4).value = 'Reduced Prices'
sheet.cell(1, 3).value = 'Prices'

for row in range(2, max + 1):
    cell = sheet.cell(row, 3)
    new_price = round(cell.value * 0.5, 2)
    # Create new row with cells to store value
    new_price_cell = sheet.cell(row, 4)
    new_price_cell.value = f'${new_price}'

work_book.save('./Python_Excel/transactions_new.xlsx')
